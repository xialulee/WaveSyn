# -*- coding: utf-8 -*-
"""
Created on Mon Jul 28 15:05:11 2014

@author: xialulee
"""
from Tkinter                         import *
from ttk                             import *
from openpyxl.reader.excel           import load_workbook
from copy                            import copy

from wavesynlib.guicomponents        import ScrolledList
from wavesynlib.languagecenter.utils import MethodDelegator

def selectSheet(workBook, ignoreBlankSheet=True, label_text='Select a sheet'):
    retval  = [0]
    
    sheetNames  = workBook.getSheetNames()
    
    if ignoreBlankSheet:
        sheetNames_backup   = copy(sheetNames)
        for name in sheetNames_backup:
            sheet   = workBook.getSheetByName(name)
            print 'rows:', sheet.get_highest_row()
            if sheet.get_highest_row() == 0:
                sheetNames.remove(name)
                
    if ignoreBlankSheet and len(sheetNames) == 1:
        return sheetNames[0]
    
    win = Toplevel()
    win.resizable(width=False, height=False)
    Label(win, text=label_text).pack()
                
    listBox = ScrolledList(win)
    listBox.pack()
    for name in sheetNames:
        listBox.insert(END, name)
        
    def onClick():
        retval[0]   = listBox.list.get(listBox.current_selection[0])
        win.destroy()
    Button(win, text='OK', command=onClick).pack()
    win.grab_set()
    win.focus_set()
    win.wait_window()
    return retval[0]
    
    
class WorkBook(object):
    method_name_map   = {
        'getSheetNames':  'get_sheet_names',
        'getSheetByName':    'get_sheet_by_name'
    }
    for method_name in method_name_map:
        locals()[method_name]    = MethodDelegator('book', method_name_map[method_name])
        
    def __init__(self, filename=None):
        if filename:
            self.load(filename)
            
    def load(self, filename):
        self.__filename = filename
        self.__book     = load_workbook(filename)
        
    def getColumn(self, sheetName, columnIdx):
        workSheet   = self.__book.get_sheet_by_name(sheetName)
        return [workSheet.cell(row=rowIdx, column=columnIdx).value for rowIdx in range(1, workSheet.get_highest_row()+1)]
        
    @property
    def book(self):
        return self.__book
            